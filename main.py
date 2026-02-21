import openpyxl
inventory_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inventory_file.active
#task1
product_per_supplier={}
total_inventory_value={}
inventory_new={}
for product_row in range(2,product_list.max_row+1):
    supplier_name=product_list.cell(product_row,4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    value = inventory * price
    product_no=product_list.cell(product_row,1).value
    new_col=product_list.cell(product_row,5)
    if supplier_name in product_per_supplier:
        current_num_products=product_per_supplier[supplier_name]
        product_per_supplier[supplier_name]=current_num_products+1
    else:
        product_per_supplier[supplier_name]=1
    if supplier_name in total_inventory_value:
        total_inventory_value[supplier_name]+=int(value)
    else:
        total_inventory_value[supplier_name]=int(value)
    if inventory<10:
        inventory_new[int(product_no)]=int(inventory)
    new_col.value=value
inventory_file.save("new1.xlsx")
print("file saved successfully")
#task1
print(product_per_supplier)
#task2
print(total_inventory_value)
#task3
print(inventory_new)


